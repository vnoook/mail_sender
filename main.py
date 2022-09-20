# Программа для рассылки HTML писем в виде HTML из XLS файла
# ...
# Содержание msc.py файла с настройками:
# --- msc file ---
# msc_mail_server = 'test.com'
# msc_login_user = 'test'
# msc_login_pass = 'pass'
# msc_from_address = 'test@test.ru'
# msc_test_address = 'test@test.ru'
# msc_subject_text = 'subject text'
# msc_flag_sending = False/True
# --- msc file ---
# ...
# INSTALL
# pip install openpyxl
# pip install PyQt5
# ...
# COMPILE
# pyinstaller -F -w main.py
# ...

import sys
import time
import smtplib
import email.mime.text
import PyQt5
import PyQt5.QtWidgets
import PyQt5.QtCore
import PyQt5.QtGui
import openpyxl
import openpyxl.utils
import msc


# класс получателя сообщения одной отправки
class RecipientData:
    """Класс Получателя сообщения"""

    # количество экземпляров класса
    count_recipient = 0

    # инициализация переменных объекта
    def __init__(self, rd_text_message=None):
        self.num = None
        self.fam = None
        self.im = None
        self.otch = None
        self.email = None
        self.mno_code = None
        self.text_message = rd_text_message
        self.flag_send_message = False

        # изменение счётчика экземпляров
        RecipientData.count_recipient += 1

    # метод получения всех значений аргументов
    def get_all_info(self):
        return f'Объект {self.get_obj_name()}, '\
               f'{id(self)}, '\
               f'{self.num = }, '\
               f'{self.fam = }, '\
               f'{self.im = }, '\
               f'{self.otch = }, '\
               f'{self.email = }, '\
               f'{self.mno_code = }, '\
               f'{self.text_message = }, '\
               f'{self.flag_send_message = }'

    # метод получения имени экземпляра
    def get_obj_name(self):
        for glob_name, glob_val in globals().items():
            if glob_val is self:
                return glob_name

    # метод замены тегов в сообщении на значения из экселя
    @staticmethod
    def replace_text_message(mail_tag, mail_tag_value, mail_text):
        mail_text = mail_text.replace(''.join(('{{', mail_tag, '}}')), mail_tag_value)
        return mail_text

    # переопределение метода для замены тегов в почтовом сообщении
    def __setattr__(self, key, value):
        if 'Recipient' in str(self.get_obj_name()):
            if key in ('num', 'fam', 'im', 'otch', 'mno_code'):
                self.text_message = self.replace_text_message(key, value, self.text_message)
        return object.__setattr__(self, key, value)


# класс потока, для отправки почты в отдельном потоке
class Thread(PyQt5.QtCore.QThread):
    """Класс отдельного потока для отправки сообщений"""

    # сигналы для прогресс-бара, информационного сообщения, окончания потока
    signal_progress_bar = PyQt5.QtCore.pyqtSignal(int)
    signal_progress_bar_setMaximum = PyQt5.QtCore.pyqtSignal(int)
    signal_actual_doing = PyQt5.QtCore.pyqtSignal(str)
    signal_finish_thread = PyQt5.QtCore.pyqtSignal()

    # инициализация класса потока
    def __init__(self, args_main_form):
        super().__init__()

        # кортеж свойств с главной формы в поток
        self.args = args_main_form
        # переменная для окна сообщения
        self.window_info = None

    # предопределённая функция запуска потока
    def run(self):
        # подсчёт времени 'начало'
        time_start = time.monotonic()

        # очищение прогресс-бара
        self.signal_progress_bar.emit(0)

        # установка текущих значений из кортежа свойств формы
        q_pocket = self.args[0]
        q_messages = self.args[1]
        send_delay = self.args[2]
        html_file = self.args[3]
        xls_file = self.args[4]
        subject_letter = self.args[5]

        # открываю файл HTML
        self.signal_actual_doing.emit(f'открытие файла HTML')
        with open(html_file, 'r') as file_html:
            all_strings_html_file = file_html.read()

        # открываю файл XLS и выбираю активный лист
        self.signal_actual_doing.emit(f'открытие файла XLS')
        wb_xls = openpyxl.load_workbook(xls_file)
        wb_xls_s = wb_xls.active

        # переменные для обработки XLS
        # список слов для замены в HTML файле
        list_replaced_words = []
        # счётчик объектов, с 0 потому что первая строка шапка и там нет обрабатываемых данных
        obj_count = 0
        # короткое обращение к объекту, утилитарная переменная
        obj_name = None

        # передача в сигналы данных для настройки прогресс-бара
        self.signal_progress_bar_setMaximum.emit(wb_xls_s.max_row)
        self.signal_progress_bar.emit(0)

        # получение значений ячеек из XLS файла
        # проход по строкам
        for row_in_xls in range(1, wb_xls_s.max_row + 1):
            # выдача информации о чтении каждой Х строки файла XLS и изменение прогресс-бара
            if row_in_xls % 25 == 0:
                self.signal_actual_doing.emit(f'получение значений ячеек из XLS файла, '
                                              f'строка {row_in_xls} из {wb_xls_s.max_row} строк')
            self.signal_progress_bar.emit(row_in_xls)

            # проход по колонкам
            for col_in_xls in range(1, wb_xls_s.max_column + 1):
                # значение ячейки
                cell_value = wb_xls_s.cell(row_in_xls, col_in_xls).value

                # если первая строка, то сформировать список спецстрок из шапки, которые нужно будет искать и заменять
                # иначе обрабатывается остальные строки с данными
                if row_in_xls == 1:
                    # считываются все, кроме email значения и вносятся для последующей теговой замены
                    if cell_value != 'email':  # если не колонка с почтами
                        list_replaced_words.append(cell_value)
                else:
                    # если первая колонка, то создаётся объект, иначе просто заполняются атрибуты из ячеек
                    if col_in_xls == 1:
                        # увеличение итерации счётчика созданных объектов
                        obj_count += 1

                        # создание объекта
                        globals()['Recipient' + str(obj_count)] = RecipientData(rd_text_message=all_strings_html_file)

                        # короткое обращение к созданному объекту
                        obj_name = globals()['Recipient' + str(obj_count)]

                        # заполнение первого аргумента
                        obj_name.__setattr__(wb_xls_s.cell(1, col_in_xls).value, str(cell_value))
                    else:
                        # заполнение остальных атрибутов по названиям колонок в верхней строке
                        obj_name.__setattr__(wb_xls_s.cell(1, col_in_xls).value, cell_value)

        # передача в сигналы данных для настройки прогресс-бара
        self.signal_progress_bar_setMaximum.emit(RecipientData.count_recipient)
        self.signal_progress_bar.emit(0)

        # участок отправки писем и ожидания времени
        # утилитарный список для хранения всех номеров объектов, будет использоваться для срезов
        list_recipients = [x for x in range(1, RecipientData.count_recipient + 1)]
        # проход по всем объектам
        for recipient in range(0, RecipientData.count_recipient, q_pocket):
            # взятие среза для выбора количества писем в пакете
            list_recipients_pocket = list_recipients[recipient: recipient + q_pocket]

            # проход по срезу, в данном случае это количество в пакете
            for recipient_number in list_recipients_pocket:
                # короткое обращение к объекту
                obj_name = globals()['Recipient' + str(recipient_number)]

                # создание текста письма
                msg = email.mime.text.MIMEText(obj_name.text_message, 'html')
                msg['From'] = msc.msc_from_address
                msg['To'] = obj_name.email
                msg['Subject'] = str(subject_letter)

                # попытка отправить письмо
                try:
                    if msc.msc_flag_sending:
                        # создание соединения с сервером
                        smtp_link = smtplib.SMTP(msc.msc_mail_server)
                        smtp_link.starttls()
                        # подключение к аккаунту
                        smtp_link.login(msc.msc_from_address, msc.msc_login_pass)
                        smtp_link.send_message(msg, msc.msc_from_address, obj_name.email)
                        smtp_link.quit()
                        obj_name.flag_send_message = True

                except Exception as _ex:
                    # информационное окно об ошибке при отправке сообщения
                    self.window_info = PyQt5.QtWidgets.QMessageBox()
                    self.window_info.setWindowTitle('Ошибка')
                    self.window_info.setText(f'Ошибка при отправке.\n{_ex}')
                    self.window_info.exec_()

                # изменения прогресс-бара
                self.signal_progress_bar.emit(recipient_number)

                # выдача информации и ожидание между письмами
                if list_recipients_pocket.index(recipient_number) != len(list_recipients_pocket) - 1:
                    self.signal_actual_doing.emit(f'отправилось {recipient_number} письмо,'
                                                  f' осталось {RecipientData.count_recipient - recipient_number} писем')
                    time.sleep(q_messages)

            # выдача информации и ожидание между пакетами
            if len(list_recipients_pocket) == q_pocket:
                if RecipientData.count_recipient not in list_recipients_pocket:
                    self.signal_actual_doing.emit(f'отправилось {recipient_number} письмо,'
                                                  f' осталось {RecipientData.count_recipient - recipient_number} писем')
                    time.sleep(send_delay)

        # выдача информации об окончании отправки
        self.signal_actual_doing.emit(f'отправка окончена')

        # очистка переменных после отправки почты
        self.clean_vals()

        # закрытие файла XLS
        wb_xls.close()

        # подсчёт времени 'конец'
        time_finish = time.monotonic()

        # информационное окно об окончании работ по отправке
        self.window_info = PyQt5.QtWidgets.QMessageBox()
        self.window_info.setWindowTitle('Отправка писем окончена')
        self.window_info.setText(f'Файлы закрыты.\n'
                                 f'Отправка писем выполнена за {round(time_finish - time_start, 1)} секунд.')
        self.window_info.exec_()

        # отправка сигнала о том, что все действия в потоке закончились
        self.signal_finish_thread.emit()

        # выдача информации об окончании отправки
        self.signal_actual_doing.emit(f'отправка окончена')

    # функция при остановке потока вручную
    def stop(self):
        self.signal_actual_doing.emit(f'отправка принудительно остановлена')
        self.signal_finish_thread.emit()
        self.terminate()

        # очистка переменных после отправки почты
        self.clean_vals()

    # метод очистки переменных после отправки почты
    @staticmethod
    def clean_vals():
        # удаление объектов отправителей
        for count_obj in range(1, RecipientData.count_recipient + 1):
            del globals()['Recipient' + str(count_obj)]

        # обнуление счётчика количества объектов для возможности повторной отправки рассылки
        RecipientData.count_recipient = 0

    # функция расчёта примерного времени требуемого для отправки всех писем
    @staticmethod
    def time_count(letters_all=16, letters_pack=5, delay_letter=3, delay_pack=300):
        q_full_pack = letters_all // letters_pack

        if letters_all % letters_pack == 0:
            time_pack = (q_full_pack - 1) * delay_pack
            time_letters = (((letters_pack - 1) * delay_letter) * q_full_pack)
            time_short_pack = 0
        else:
            time_pack = q_full_pack * delay_pack
            time_letters = (((letters_pack - 1) * delay_letter) * q_full_pack)
            time_short_pack = ((letters_all % letters_pack) - 1) * delay_letter

        time_all = time_pack + time_letters + time_short_pack
        return time_all


# класс главного окна
class Window(PyQt5.QtWidgets.QMainWindow):
    """Класс главного окна"""

    # описание главного окна
    def __init__(self):
        super().__init__()

        # словарь для хранения потока
        self.thread = {}

        # переменные
        self.window_info = None
        self.info_for_open_file = None
        self.info_path_open_file = None

        self.info_extention_open_file_html = 'Файлы HTML (*.html; *.htm)'
        self.info_extention_open_file_xls = 'Файлы Excel xlsx (*.xlsx)'
        self.text_empty_path_file = 'файл пока не выбран'

        # количество писем в одном пакете отправки, в штуках
        self.q_pocket = 5
        # задержка между письмами в пакете при отправке, в секундах
        self.q_messages = 3
        # задержка между отправками пакетов, в секундах
        self.send_delay = 300  # 5 минут

        # главное окно, надпись на нём и размеры
        self.setWindowTitle('Рассылка почты из XLS файла на основе шаблона HTML')
        self.setGeometry(450, 100, 700, 490)

        # ОБЪЕКТЫ НА ФОРМЕ
        # HTML
        # label_html_file
        self.label_html_file = PyQt5.QtWidgets.QLabel(self)
        self.label_html_file.setObjectName('label_html_file')
        self.label_html_file.setText('1. Выберите HTML файл шаблона')
        self.label_html_file.setGeometry(PyQt5.QtCore.QRect(10, 10, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_html_file.setFont(font)
        self.label_html_file.adjustSize()
        self.label_html_file.setToolTip(self.label_html_file.objectName())

        # toolButton_select_html_file
        self.toolButton_select_html_file = PyQt5.QtWidgets.QPushButton(self)
        self.toolButton_select_html_file.setObjectName('toolButton_select_html_file')
        self.toolButton_select_html_file.setText('...')
        self.toolButton_select_html_file.setGeometry(PyQt5.QtCore.QRect(10, 40, 50, 20))
        self.toolButton_select_html_file.setFixedWidth(50)
        self.toolButton_select_html_file.clicked.connect(self.select_file)
        self.toolButton_select_html_file.setToolTip(self.toolButton_select_html_file.objectName())

        # label_path_html_file
        self.label_path_html_file = PyQt5.QtWidgets.QLabel(self)
        self.label_path_html_file.setObjectName('label_path_html_file')
        self.label_path_html_file.setText(self.text_empty_path_file)
        self.label_path_html_file.setGeometry(PyQt5.QtCore.QRect(70, 40, 820, 16))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(10)
        self.label_path_html_file.setFont(font)
        self.label_path_html_file.adjustSize()
        self.label_path_html_file.setToolTip(self.label_path_html_file.objectName())

        # XLS
        # label_xls_file
        self.label_xls_file = PyQt5.QtWidgets.QLabel(self)
        self.label_xls_file.setObjectName('label_xls_file')
        self.label_xls_file.setText('2. Выберите EXCEL файл со справочником адресатов')
        self.label_xls_file.setGeometry(PyQt5.QtCore.QRect(10, 70, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_xls_file.setFont(font)
        self.label_xls_file.adjustSize()
        self.label_xls_file.setToolTip(self.label_xls_file.objectName())

        # toolButton_select_xls_file
        self.toolButton_select_xls_file = PyQt5.QtWidgets.QPushButton(self)
        self.toolButton_select_xls_file.setObjectName('toolButton_select_xls_file')
        self.toolButton_select_xls_file.setText('...')
        self.toolButton_select_xls_file.setGeometry(PyQt5.QtCore.QRect(10, 100, 50, 20))
        self.toolButton_select_xls_file.setFixedWidth(50)
        self.toolButton_select_xls_file.clicked.connect(self.select_file)
        self.toolButton_select_xls_file.setToolTip(self.toolButton_select_xls_file.objectName())

        # label_path_xls_file
        self.label_path_xls_file = PyQt5.QtWidgets.QLabel(self)
        self.label_path_xls_file.setObjectName('label_path_xls_file')
        self.label_path_xls_file.setText(self.text_empty_path_file)
        self.label_path_xls_file.setGeometry(PyQt5.QtCore.QRect(70, 100, 820, 20))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(10)
        self.label_path_xls_file.setFont(font)
        self.label_path_xls_file.adjustSize()
        self.label_path_xls_file.setToolTip(self.label_path_xls_file.objectName())

        # Q_POCKET
        # label_q_pocket
        self.label_q_pocket = PyQt5.QtWidgets.QLabel(self)
        self.label_q_pocket.setObjectName('label_q_pocket')
        self.label_q_pocket.setText('3. Сколько писем в одном пакете, шт.')
        self.label_q_pocket.setGeometry(PyQt5.QtCore.QRect(10, 130, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_q_pocket.setFont(font)
        self.label_q_pocket.adjustSize()
        self.label_q_pocket.setToolTip(self.label_q_pocket.objectName())

        # lineEdit_q_pocket
        self.lineEdit_q_pocket = PyQt5.QtWidgets.QLineEdit(self)
        self.lineEdit_q_pocket.setObjectName('lineEdit_q_pocket')
        self.lineEdit_q_pocket.setPlaceholderText('Введите число')
        self.lineEdit_q_pocket.setText(str(self.q_pocket))
        self.lineEdit_q_pocket.setGeometry(PyQt5.QtCore.QRect(10, 160, 110, 20))
        self.lineEdit_q_pocket.setClearButtonEnabled(True)
        self.lineEdit_q_pocket.setEnabled(False)
        self.lineEdit_q_pocket.setToolTip(self.lineEdit_q_pocket.objectName())

        # Q_MESSAGES
        # label_q_messages
        self.label_q_messages = PyQt5.QtWidgets.QLabel(self)
        self.label_q_messages.setObjectName('label_q_messages')
        self.label_q_messages.setText('4. Задержка между письмами, сек.')
        self.label_q_messages.setGeometry(PyQt5.QtCore.QRect(10, 190, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_q_messages.setFont(font)
        self.label_q_messages.adjustSize()
        self.label_q_messages.setToolTip(self.label_q_messages.objectName())

        # lineEdit_q_messages
        self.lineEdit_q_messages = PyQt5.QtWidgets.QLineEdit(self)
        self.lineEdit_q_messages.setObjectName('lineEdit_q_messages')
        self.lineEdit_q_messages.setPlaceholderText('Введите число')
        self.lineEdit_q_messages.setText(str(self.q_messages))
        self.lineEdit_q_messages.setGeometry(PyQt5.QtCore.QRect(10, 220, 110, 20))
        self.lineEdit_q_messages.setClearButtonEnabled(True)
        self.lineEdit_q_messages.setEnabled(False)
        self.lineEdit_q_messages.setToolTip(self.lineEdit_q_messages.objectName())

        # MAIL_DELAY
        # label_mail_delay
        self.label_mail_delay = PyQt5.QtWidgets.QLabel(self)
        self.label_mail_delay.setObjectName('label_mail_delay')
        self.label_mail_delay.setText('5. Задержка между отправками пакетов, сек.')
        self.label_mail_delay.setGeometry(PyQt5.QtCore.QRect(10, 250, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_mail_delay.setFont(font)
        self.label_mail_delay.adjustSize()
        self.label_mail_delay.setToolTip(self.label_mail_delay.objectName())

        # lineEdit_mail_delay
        self.lineEdit_mail_delay = PyQt5.QtWidgets.QLineEdit(self)
        self.lineEdit_mail_delay.setObjectName('lineEdit_mail_delay')
        self.lineEdit_mail_delay.setPlaceholderText('Введите число')
        self.lineEdit_mail_delay.setText(str(self.send_delay))
        self.lineEdit_mail_delay.setGeometry(PyQt5.QtCore.QRect(10, 280, 110, 25))
        self.lineEdit_mail_delay.setClearButtonEnabled(True)
        self.lineEdit_mail_delay.setEnabled(False)
        self.lineEdit_mail_delay.setToolTip(self.lineEdit_mail_delay.objectName())

        # lineEdit_subject_letter
        self.lineEdit_subject_letter = PyQt5.QtWidgets.QLineEdit(self)
        self.lineEdit_subject_letter.setObjectName('lineEdit_subject_letter')
        self.lineEdit_subject_letter.setPlaceholderText('Введите тему письма!')
        self.lineEdit_subject_letter.setGeometry(PyQt5.QtCore.QRect(10, 320, 320, 25))
        self.lineEdit_subject_letter.setClearButtonEnabled(True)
        self.lineEdit_subject_letter.setEnabled(False)
        self.lineEdit_subject_letter.setToolTip(self.lineEdit_subject_letter.objectName())

        # SEND_MAIL
        # pushButton_send_mail
        self.pushButton_send_mail = PyQt5.QtWidgets.QPushButton(self)
        self.pushButton_send_mail.setObjectName('pushButton_send_mail')
        self.pushButton_send_mail.setEnabled(False)
        self.pushButton_send_mail.setText('Отправьте почту')
        self.pushButton_send_mail.setGeometry(PyQt5.QtCore.QRect(10, 360, 180, 25))
        self.pushButton_send_mail.setFixedWidth(130)
        self.pushButton_send_mail.clicked.connect(self.init_thread)
        self.pushButton_send_mail.setToolTip(self.pushButton_send_mail.objectName())

        # SEND_TEST_MAIL
        # pushButton_send_test_mail
        self.pushButton_send_test_mail = PyQt5.QtWidgets.QPushButton(self)
        self.pushButton_send_test_mail.setObjectName('pushButton_send_test_mail')
        self.pushButton_send_test_mail.setEnabled(False)
        self.pushButton_send_test_mail.setText('Тестовое письмо')
        self.pushButton_send_test_mail.setGeometry(PyQt5.QtCore.QRect(200, 360, 300, 25))
        self.pushButton_send_test_mail.setFixedWidth(130)
        self.pushButton_send_test_mail.clicked.connect(self.send_test_mail)
        self.pushButton_send_test_mail.setToolTip(self.pushButton_send_test_mail.objectName())

        # STATISTICS
        # progressBarStat
        self.progressBarStat = PyQt5.QtWidgets.QProgressBar(self)
        self.progressBarStat.setObjectName('progressBarStat')
        self.progressBarStat.setGeometry(PyQt5.QtCore.QRect(10, 400, 320, 25))
        self.progressBarStat.setMinimum(0)
        self.progressBarStat.setMaximum(100)
        self.progressBarStat.setValue(0)
        self.progressBarStat.setFormat('%p %')
        self.progressBarStat.setToolTip(self.progressBarStat.objectName())

        # EXIT
        # button_exit
        self.button_exit = PyQt5.QtWidgets.QPushButton(self)
        self.button_exit.setObjectName('button_exit')
        self.button_exit.setText('Выход')
        self.button_exit.setGeometry(PyQt5.QtCore.QRect(10, 450, 180, 25))
        self.button_exit.setFixedWidth(50)
        self.button_exit.clicked.connect(self.click_on_btn_exit)
        self.button_exit.setToolTip(self.button_exit.objectName())

        # label_info
        self.label_info = PyQt5.QtWidgets.QLabel(self)
        self.label_info.setObjectName('label_info')
        self.label_info.clear()
        self.label_info.setGeometry(PyQt5.QtCore.QRect(70, 455, 300, 25))
        self.label_info.setEnabled(False)
        self.label_info.adjustSize()
        self.label_info.setToolTip(self.label_info.objectName())

        # INVISIBLE
        # checkBox_inviz
        self.checkBox_inviz = PyQt5.QtWidgets.QCheckBox(self)
        self.checkBox_inviz.setObjectName('checkBox_inviz')
        self.checkBox_inviz.setGeometry(PyQt5.QtCore.QRect(10, 500, 190, 40))
        self.checkBox_inviz.clicked.connect(self.on_off_lineedits_delays)
        self.checkBox_inviz.setText('Хочу редактировать!')
        self.checkBox_inviz.setToolTip(self.checkBox_inviz.objectName())

    # метод инициализации 'что нужно делать' с потоком - стартовать или останавливать
    def init_thread(self):
        # выбор функции зависит от пустоты словаря
        # если там есть поток, то надо стопнуть
        # если словарь пустой, то запустить
        if not self.thread:
            # проверка полей формы на правильность заполнения и выбор действия
            if self.check_fields():
                self.start_thread()
            else:
                # информационное окно, очистка прогресс-бара и информационной строки
                PyQt5.QtWidgets.QMessageBox.information(self, 'Внимание', 'Заполните все поля правильно!')
                self.change_progressbarstat_val(0)
                self.label_info.clear()
        else:
            # принудительная остановка потока
            self.stop_thread()

    # подготовка к созданию потока, проверка всех полей на форме
    def check_fields(self):
        # флаг правильности заполненных полей
        flag_check = False

        # объекты, которые проверяются
        q_mes = self.lineEdit_q_messages
        q_poc = self.lineEdit_q_pocket
        m_del = self.lineEdit_mail_delay
        s_let = self.lineEdit_subject_letter

        # кортеж полей с числами и проверка на введённое число,
        # если не число, то поле становится пустым
        tuple_fields__for_digits = (q_mes, q_poc, m_del)
        for obj in tuple_fields__for_digits:
            if self.check_is_digit(obj.text()):
                flag_check = True
            else:
                flag_check = False
                obj.clear()

        # кортеж проверяемых данных из полей объектов на форме и проверка их на пустоту
        # тема письма может быть только непустой,
        # остальные поля могут быть заполненными, но заполнены неправильными данными
        tuple_of_fields = (self.check_is_digit(q_mes.text()),
                           self.check_is_digit(q_poc.text()),
                           self.check_is_digit(m_del.text()),
                           s_let.text())
        if all(tuple_of_fields):
            flag_check = True
        else:
            flag_check = False

        # установка итогового результата исходя из значения флага
        if flag_check:
            return True
        else:
            return False

    # метод старта потока и привязка сигналов к функциям
    def start_thread(self):
        # кортеж для передачи задержек в объект потока
        # тут смелый int потому, что я выше проверил значения
        args_main_form = (int(self.lineEdit_q_pocket.text()),
                          int(self.lineEdit_q_messages.text()),
                          int(self.lineEdit_mail_delay.text()),
                          self.label_path_html_file.text(),
                          self.label_path_xls_file.text(),
                          self.lineEdit_subject_letter.text(),
                          )

        # создание объекта потока
        self.thread['Thread'] = Thread(args_main_form)
        self.thread['Thread'].start()
        self.thread['Thread'].signal_actual_doing.connect(self.show_actual_doing)
        self.thread['Thread'].signal_progress_bar.connect(self.change_progressbarstat_val)
        self.thread['Thread'].signal_progress_bar_setMaximum.connect(self.change_progressbarstat_set_maximum)
        self.thread['Thread'].signal_finish_thread.connect(self.finished)

        # деактивация объектов на форме
        self.activate_obj_on_form(0)

    # метод остановки потока и обнуление словаря потока
    def stop_thread(self):
        # останавливаю работающий поток
        self.thread['Thread'].stop()
        # очищаю словарь потоков
        self.thread = {}
        # активация объектов на форме
        self.activate_obj_on_form(1)

    # метод должен успешного окончания потока
    def finished(self):
        # очистка словаря хранения потока
        self.thread = {}
        # активация объектов на форме
        self.activate_obj_on_form(1)

    # метод для изменения максимального значения прогресс-бара на форме
    def change_progressbarstat_set_maximum(self, val_int):
        self.progressBarStat.setMaximum(val_int)

    # метод для изменения прогресс-бара на форме
    def change_progressbarstat_val(self, val_int):
        self.progressBarStat.setValue(val_int)

    # метод для передачи информации в информационную строку
    def show_actual_doing(self, string_actual_doing):
        self.label_info.setText(string_actual_doing)
        self.label_info.adjustSize()

    # событие - скрытие\отображение возможности редактирования полей задержек при отправке
    def on_off_lineedits_delays(self):
        if self.checkBox_inviz.isChecked():
            self.lineEdit_q_pocket.setEnabled(True)
            self.lineEdit_q_messages.setEnabled(True)
            self.lineEdit_mail_delay.setEnabled(True)
        else:
            self.lineEdit_q_pocket.setEnabled(False)
            self.lineEdit_q_messages.setEnabled(False)
            self.lineEdit_mail_delay.setEnabled(False)

    # событие - скрытие\отображение возможности редактирования полей во время отправки
    def activate_obj_on_form(self, action_todo):
        if action_todo == 0:
            self.toolButton_select_html_file.setEnabled(False)
            self.toolButton_select_xls_file.setEnabled(False)
            self.lineEdit_q_pocket.setEnabled(False)
            self.lineEdit_q_messages.setEnabled(False)
            self.lineEdit_mail_delay.setEnabled(False)
            self.pushButton_send_test_mail.setEnabled(False)
            self.checkBox_inviz.setEnabled(False)
            self.lineEdit_subject_letter.setEnabled(False)
            self.pushButton_send_mail.setText('Прекратить отправку')
        elif action_todo == 1:
            self.toolButton_select_html_file.setEnabled(True)
            self.toolButton_select_xls_file.setEnabled(True)
            self.pushButton_send_test_mail.setEnabled(True)
            self.checkBox_inviz.setChecked(False)
            self.checkBox_inviz.setEnabled(True)
            self.lineEdit_subject_letter.setEnabled(True)
            self.pushButton_send_mail.setText('Отправьте почту')

    # событие - нажатие на кнопку выбора файла
    def select_file(self):
        # переменная для хранения информации из окна выбора файла
        data_of_open_file_name = None
        # запоминание старого значения пути выбора файлов
        old_path_of_selected_html_file = self.label_path_html_file.text()
        old_path_of_selected_xls_file = self.label_path_xls_file.text()

        # определение какая кнопка выбора файла нажата
        if self.sender().objectName() == self.toolButton_select_html_file.objectName():
            self.info_for_open_file = 'Выберите HTML файл (.HTML или .HTM)'
            # непосредственное окно выбора файла и переменная для хранения пути файла
            data_of_open_file_name = PyQt5.QtWidgets.QFileDialog.getOpenFileName(self,
                                                                                 self.info_for_open_file,
                                                                                 self.info_path_open_file,
                                                                                 self.info_extention_open_file_html)
        elif self.sender().objectName() == self.toolButton_select_xls_file.objectName():
            self.info_for_open_file = 'Выберите Excel (.XLSX)'
            # непосредственное окно выбора файла и переменная для хранения пути файла
            data_of_open_file_name = PyQt5.QtWidgets.QFileDialog.getOpenFileName(self,
                                                                                 self.info_for_open_file,
                                                                                 self.info_path_open_file,
                                                                                 self.info_extention_open_file_xls)

        # выбор только пути файла из data_of_open_file_name
        file_name = data_of_open_file_name[0]

        # выбор где и что менять исходя из выбора пользователя
        # нажата кнопка выбора HTML файла
        if self.sender().objectName() == self.toolButton_select_html_file.objectName():
            if file_name == '':
                self.label_path_html_file.setText(old_path_of_selected_html_file)
                self.label_path_html_file.adjustSize()
            else:
                old_path_of_selected_html_file = self.label_path_html_file.text()
                self.label_path_html_file.setText(file_name)
                self.label_path_html_file.adjustSize()

        # нажата кнопка выбора XLS файла
        if self.sender().objectName() == self.toolButton_select_xls_file.objectName():
            if file_name == '':
                self.label_path_xls_file.setText(old_path_of_selected_xls_file)
                self.label_path_xls_file.adjustSize()
            else:
                old_path_of_selected_xls_file = self.label_path_xls_file.text()
                self.label_path_xls_file.setText(file_name)
                self.label_path_xls_file.adjustSize()

        # активация и деактивация объектов на форме зависящее от 'выбраны ли все файлы' и 'они разные'
        if self.text_empty_path_file not in (self.label_path_html_file.text(), self.label_path_xls_file.text()):
            self.pushButton_send_mail.setEnabled(True)
            self.pushButton_send_test_mail.setEnabled(True)
            self.lineEdit_subject_letter.setEnabled(True)

    # событие - нажатие на кнопку отправки тестового письма
    def send_test_mail(self):
        # открываю и читаю файл HTML
        with open(self.label_path_html_file.text(), 'r') as file_html:
            all_strings_html_file = file_html.read()

        # создание соединения с сервером
        smtp_link = smtplib.SMTP(msc.msc_mail_server)
        smtp_link.starttls()

        try:
            # подключение к аккаунту
            smtp_link.login(msc.msc_from_address, msc.msc_login_pass)
            # создание текста письма
            msg = email.mime.text.MIMEText(all_strings_html_file, 'html')
            msg['From'] = msc.msc_from_address
            msg['To'] = msc.msc_test_address
            msg['Subject'] = msc.msc_subject_text
            if msc.msc_flag_sending:
                smtp_link.send_message(msg, msc.msc_from_address, msc.msc_test_address)
            smtp_link.quit()

            # информационное окно об удачной отправке тестового письма
            PyQt5.QtWidgets.QMessageBox.information(self, 'Отправлено', f'Тестовое письмо отправлено '
                                                                        f'на почту {msc.msc_test_address}.')

        except Exception as _ex:
            # информационное окно об ошибке при отправке сообщения
            PyQt5.QtWidgets.QMessageBox.information(self, 'Ошибка', f'Ошибка при отправке.\n{_ex}')

    # проверка строки на числовое значение в строке на форме
    # число в строке должно быть больше 0, потому что в этих полях вводятся задержки в секундах
    @staticmethod
    def check_is_digit(data):
        try:
            data_int = int(data)
            if data_int > 0:
                return True
            else:
                return False
        except ValueError:
            return False

    # событие - нажатие на кнопку Выход
    @staticmethod
    def click_on_btn_exit():
        sys.exit()


# создание основного окна
def main_app():
    app = PyQt5.QtWidgets.QApplication(sys.argv)
    app.setStyle('Fusion')
    app_window_main = Window()
    app_window_main.show()
    sys.exit(app.exec_())


# запуск основного окна
if __name__ == '__main__':
    main_app()
